import io
from openpyxl import Workbook
from openpyxl.styles import Font


def build_bookings_report(bookings):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings"

    headers = [
        "Reference", "Attendee Name", "Email", "Phone", "Seats",
        "Amount (EUR)", "Payment Method", "Payment Status", "Booked At", "Checked In",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for b in bookings:
        checkin_detail = "; ".join(
            f"{s.label}: {s.checked_in_at.strftime('%Y-%m-%d %H:%M') if s.checked_in_at else 'not checked in'}"
            for s in sorted(b.seats, key=lambda s: s.label)
        )
        ws.append([
            b.reference,
            b.attendee_name,
            b.email,
            b.phone or "",
            b.seat_labels(),
            b.amount_total_eur,
            b.payment_method,
            b.payment_status,
            b.created_at.strftime("%Y-%m-%d %H:%M") if b.created_at else "",
            checkin_detail,
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_analytics_report(summary, daily_rows):
    wb = Workbook()

    ws = wb.active
    ws.title = "Funnel"
    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(["Page views", summary["page_views"]])
    ws.append(["Visits (unique sessions)", summary["visits"]])
    ws.append(["Clicked \"Book Your Seat\"", summary["book_clicks"]])
    ws.append(["Selected a seat", summary["seat_selected"]])
    ws.append(["Started checkout (Stripe)", summary["checkout_started"]])
    ws.append(["Booked online", summary["booked_online"]])
    ws.append(["Booked for cash (admin)", summary["booked_cash"]])
    ws.append(["Booked total", summary["booked_total"]])
    ws.append([])
    ws.append(["Conversion rate", "%"])
    ws.append(["Visit -> clicked Book", summary["rate_visit_to_click"]])
    ws.append(["Clicked Book -> selected a seat", summary["rate_click_to_seat"]])
    ws.append(["Selected a seat -> started checkout", summary["rate_seat_to_checkout"]])
    ws.append(["Started checkout -> booked online", summary["rate_checkout_to_booked"]])
    ws.append(["Visit -> booked online (overall)", summary["rate_visit_to_booked"]])
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    ws2 = wb.create_sheet("By day")
    ws2.append(["Date", "Visits", "Booked online", "Booked cash"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for row in daily_rows:
        ws2.append([row["date"].strftime("%Y-%m-%d"), row["visits"], row["booked_online"], row["booked_cash"]])
    for col in ws2.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
